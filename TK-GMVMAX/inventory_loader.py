"""inventory_loader.py - 读取 KCXQ 文件夹中的最新库存 Excel 并生成面板数据。
数据文件按日期命名，放在 KCXQ 文件夹中；每次覆盖/更新后无需重启，刷新页面即可。
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

from common import app_dir


BASE_DIR = app_dir()

# 库存数据文件夹：面板目录下的 KCXQ
KCXQ_DIRS = [
    BASE_DIR / "KCXQ",
]

_cache: dict = {"sig": None, "payload": None}

# 内置品牌关键词：Goods Name 中出现关键词即归入对应品牌，未匹配归入「其他」。
# 可在 KCXQ 目录放 brand_keywords.json 覆盖（{"品牌名": ["关键词", ...]}），无需改代码。
DEFAULT_BRAND_KEYWORDS = {"Solestorm": ["solestorm"], "Zorwalk": ["zorwalk"]}
BRAND_CONFIG_NAME = "brand_keywords.json"

# 库存账龄分桶：SKU 级 >90 天值由 91 天以上各分桶求和，保证前端账龄图与汇总口径一致。
AGING_BUCKETS = [
    "Inventory Aged 0-30 Days",
    "Inventory Aged 31-60 Days",
    "Inventory Aged 61-90 Days",
    "Inventory Aged 91-120 Days",
    "Inventory Aged 121-180 Days",
    "Inventory Aged 181-270 Days",
    "Inventory Aged 271-365 Days",
    "Inventory Aged over 365 Days",
]
AGING_OVER90_BUCKETS = AGING_BUCKETS[3:]

# 库存状态别名：TikTok 导出语言变化时统一归一到面板内部中文状态。
INVENTORY_STATUS_MAP = {
    "缺货": "缺货", "out of stock": "缺货", "out_of_stock": "缺货", "oos": "缺货",
    "即将缺货": "即将缺货", "low stock": "即将缺货", "low_stock": "即将缺货",
    "replenish soon": "即将缺货", "replenishing": "即将缺货",
    "健康": "健康", "healthy": "健康", "in stock": "健康", "normal": "健康",
}

# 产品代码提取：SKU 引用码开头的“字母+数字”段（如 SZW011-BLK-RED → SZW011）。
PRODUCT_CODE_PATTERN = re.compile(r"^([A-Za-z]{2,}\d+)")


def _normalize_status(value) -> str:
    """把中英文库存状态归一到 健康 / 缺货 / 即将缺货，未知状态保持原值。"""
    text = str(value or "").strip()
    return INVENTORY_STATUS_MAP.get(text.lower(), text or "未知")


def _extract_product_code(reference_code) -> str:
    """从 Goods Reference Code 提取产品代码；无法识别时返回 UNKNOWN。"""
    code = str(reference_code or "").strip()
    match = PRODUCT_CODE_PATTERN.match(code)
    if match:
        return match.group(1)[:20]
    return "UNKNOWN"


def _load_brand_keywords() -> tuple[dict[str, list[str]], str | None]:
    """读取可选品牌配置；无配置/配置损坏时回退内置关键词并返回警告。"""
    for directory in KCXQ_DIRS:
        path = directory / BRAND_CONFIG_NAME
        if not path.exists():
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:  # noqa: BLE001 - 配置错误回退默认即可
            return DEFAULT_BRAND_KEYWORDS, f"无法读取 {path.name}: {exc}，已使用内置品牌关键词。"
        if not isinstance(data, dict):
            return DEFAULT_BRAND_KEYWORDS, f"{path.name} 格式无效（应为 {{品牌: [关键词]}}），已使用内置品牌关键词。"
        cleaned: dict[str, list[str]] = {}
        for brand, words in data.items():
            if not isinstance(words, list):
                continue
            keywords = [str(w).strip().lower() for w in words if str(w).strip()]
            if brand and keywords:
                cleaned[str(brand)] = keywords
        return cleaned or DEFAULT_BRAND_KEYWORDS, None
    return DEFAULT_BRAND_KEYWORDS, None


def _detect_brand(goods_name: str, keywords: dict[str, list[str]]) -> str:
    """按品牌关键词从商品名识别品牌，未匹配返回「其他」。"""
    nm = (goods_name or "").lower()
    for brand, words in keywords.items():
        if any(word in nm for word in words):
            return brand
    return "其他"


def _candidate_files() -> list[Path]:
    """返回所有候选 KCXQ 目录下的 xlsx 文件（跳过 Excel 临时锁文件），按修改时间最新优先。"""
    files: list[Path] = []
    for d in KCXQ_DIRS:
        if d.exists():
            files += [p for p in d.glob("*.xlsx") if not p.name.startswith("~$")]
    files.sort(key=lambda p: p.stat().st_mtime_ns, reverse=True)
    return files


def inventory_signature() -> str:
    files = _candidate_files()
    if not files:
        return "none"
    f = files[0]
    return f"{f.name}:{f.stat().st_mtime_ns}:{f.stat().st_size}"


def _num(v) -> float:
    if v is None or v == "-" or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def _aging_over_90(row: dict) -> float:
    """>90 天库龄 = 91 天以上各分桶之和；分桶全为 0 时回退专用合计列（老导出兼容）。"""
    total = sum(_num(row.get(bucket)) for bucket in AGING_OVER90_BUCKETS)
    if total > 0:
        return total
    return _num(row.get("Inventory Aged over 90 Days"))


def _is_blank_row(values: tuple) -> bool:
    """openpyxl 会把带格式的空行也读进来：全空行直接跳过，避免幽灵 SKU。"""
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _parse_file(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_blank_row(row):
            continue
        r = {}
        for i, h in enumerate(headers):
            r[h] = row[i] if i < len(row) else None
        data.append(r)

    summary: dict = {}

    statuses: dict = {}
    for r in data:
        s = _normalize_status(r.get("Inventory Status"))
        statuses[s] = statuses.get(s, 0) + 1
    summary["inventory_status"] = statuses

    tags: dict = {}
    for r in data:
        t = r.get("Goods Tag") or "未知"
        tags[t] = tags.get(t, 0) + 1
    summary["goods_tags"] = tags

    avail = sum(_num(r.get("Available Inventory (units)")) for r in data)
    transit = sum(_num(r.get("In-Transit: Total Inventory (units)")) for r in data)
    reserved = sum(_num(r.get("Reserved (units)")) for r in data)
    excess = sum(_num(r.get("Excess Units")) for r in data)
    sales30 = sum(_num(r.get("Last 30 days sales")) for r in data)

    summary["totals"] = {
        "available": avail,
        "transit": transit,
        "reserved": reserved,
        "excess": excess,
        "total_skus": len(data),
        "total_sales_30d": sales30,
    }
    summary["sales"] = {
        "total_30d": sales30,
        "skus_with_data": len([r for r in data if _num(r.get("Last 30 days sales")) > 0]),
    }

    aging = {}
    for b in AGING_BUCKETS:
        total = sum(_num(r.get(b)) for r in data)
        aging[b.replace("Inventory Aged ", "").replace(" Days", "")] = total
    summary["aging"] = aging

    dos_dist = {"zero": 0, "low_under30": 0, "healthy_30to90": 0, "overstock_over90": 0}
    for r in data:
        v = r.get("Days of Supply (next 30 days)")
        if v is None or v == "-" or v == "":
            continue
        try:
            d = float(str(v))
            if d == 0:
                dos_dist["zero"] += 1
            elif d < 30:
                dos_dist["low_under30"] += 1
            elif d < 90:
                dos_dist["healthy_30to90"] += 1
            else:
                dos_dist["overstock_over90"] += 1
        except (TypeError, ValueError):
            pass
    summary["dos_distribution"] = dos_dist

    vol_dist = {"0": 0, "1-10": 0, "11-50": 0, "51-100": 0, "100+": 0}
    for r in data:
        v = _num(r.get("Last 30 days sales"))
        if v == 0:
            vol_dist["0"] += 1
        elif v <= 10:
            vol_dist["1-10"] += 1
        elif v <= 50:
            vol_dist["11-50"] += 1
        elif v <= 100:
            vol_dist["51-100"] += 1
        else:
            vol_dist["100+"] += 1
    summary["sales_volume"] = vol_dist

    brand_keywords, brand_warning = _load_brand_keywords()
    brands = {}
    for r in data:
        brand = _detect_brand(str(r.get("Goods Name", "") or ""), brand_keywords)
        if brand not in brands:
            brands[brand] = {"skus": 0, "available": 0, "transit": 0, "sales_30d": 0, "oos": 0}
        brands[brand]["skus"] += 1
        brands[brand]["available"] += _num(r.get("Available Inventory (units)"))
        brands[brand]["transit"] += _num(r.get("In-Transit: Total Inventory (units)"))
        brands[brand]["sales_30d"] += _num(r.get("Last 30 days sales"))
        if _normalize_status(r.get("Inventory Status")) == "缺货":
            brands[brand]["oos"] += 1
    summary["brands"] = brands

    warnings: list[str] = []
    if brand_warning:
        warnings.append(brand_warning)
    summary["warnings"] = warnings

    unknown_codes = 0
    sku_list = []
    for r in data:
        pc = _extract_product_code(r.get("Goods Reference Code", ""))
        if pc == "UNKNOWN" and str(r.get("Goods Reference Code", "") or "").strip():
            unknown_codes += 1
        sku_list.append({
            "goods_id": r.get("Goods ID", ""),
            "product_code": pc,
            "sku": r.get("Goods Reference Code", ""),
            "name": r.get("Goods Name", ""),
            "available": _num(r.get("Available Inventory (units)")),
            "transit": _num(r.get("In-Transit: Total Inventory (units)")),
            "reserved": _num(r.get("Reserved (units)")),
            "sales_30d": _num(r.get("Last 30 days sales")),
            "sales_60d": _num(r.get("Last 60 days sales")),
            "sales_90d": _num(r.get("Last 90 days sales")),
            "sell_through": _num(r.get("Sell-through Rate (last 30 days)")),
            "status": _normalize_status(r.get("Inventory Status")),
            "dos_30d": _num(r.get("Days of Supply (next 30 days)")),
            "demand_30d": _num(r.get("Demand Forecast (next 30 days)")),
            "datel_avail_30d": _num(r.get("Days of Total Stock Left (next 30 days)")),
            "aging_0_30": _num(r.get("Inventory Aged 0-30 Days")),
            "aging_31_60": _num(r.get("Inventory Aged 31-60 Days")),
            "aging_61_90": _num(r.get("Inventory Aged 61-90 Days")),
            "aging_over_90": _aging_over_90(r),
            "replenishment_date": r.get("Replenishment Ship Date", ""),
            "replenishment_qty": r.get("Replenishment Quantity", ""),
        })
    if unknown_codes:
        warnings.append(f"有 {unknown_codes} 个 SKU 的产品代码无法识别，已归入 UNKNOWN。")
    summary["skus"] = sku_list
    summary["top_skus"] = sorted(sku_list, key=lambda x: -x["sales_30d"])[:30]
    return summary


def load_inventory() -> dict:
    """返回最新库存数据（带缓存，源文件变化时自动重读）。找不到文件时返回带 error 的结构。"""
    files = _candidate_files()
    if not files:
        return {"meta": {"error": "没有在 KCXQ 文件夹中找到库存 Excel 文件。"}}
    f = files[0]
    sig = inventory_signature()
    if _cache["sig"] != sig or _cache["payload"] is None:
        _cache["sig"] = sig
        _cache["payload"] = _parse_file(f)
    payload = dict(_cache["payload"])
    payload["meta"] = {
        "file": f.name,
        "updated": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "rows": len(payload.get("skus", [])),
    }
    return payload


if __name__ == "__main__":
    import json
    result = load_inventory()
    print(json.dumps(result.get("meta", {}), ensure_ascii=False))
    print("totals:", json.dumps(result.get("totals", {}), ensure_ascii=False))
